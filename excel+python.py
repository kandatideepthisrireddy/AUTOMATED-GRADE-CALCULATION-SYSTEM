import pandas as pd
import random
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.chart import BarChart, PieChart, ScatterChart, LineChart, Reference, Series

# Function to calculate grade
def calculate_grade(average):
    if average >= 90:
        return 'A'
    elif average >= 80:
        return 'B'
    elif average >= 70:
        return 'C'
    elif average >= 60:
        return 'D'
    else:
        return 'F'

# Generate student data
student_names = [f'Student {i + 1}' for i in range(20)]
data = []
for i in range(5):
    scores = [random.randint(90, 100) for _ in range(6)]
    data.append([student_names[i]] + scores)

# Generate scores for the remaining students
for name in student_names[5:]:
    scores = [random.randint(50, 100) for _ in range(6)]
    data.append([name] + scores)

# Define columns
columns = ['Student Name', 'Math', 'Science', 'English', 'History', 'Chemistry', 'Physics']

# Create a DataFrame
df = pd.DataFrame(data, columns=columns)

# Calculate Total Marks, Average Marks, and Grade
df['Total Marks'] = df[['Math', 'Science', 'English', 'History', 'Chemistry', 'Physics']].sum(axis=1)
df['Average Marks'] = df['Total Marks'] / 6
df['Grade'] = df['Average Marks'].apply(calculate_grade)

# Save the data to an Excel file
file_path = 'student_grades.xlsx'
df.to_excel(file_path, index=False)

# Create a dashboard using openpyxl
wb = load_workbook(file_path)
ws = wb.active

# Add a summary section
summary_start_row = len(df) + 4
ws[f"A{summary_start_row}"] = "Summary"
ws[f"A{summary_start_row}"].font = Font(bold=True, size=14)

# Grade distribution
grade_distribution = df['Grade'].value_counts()
total_students = len(df)
grade_row = summary_start_row + 2

ws[f"A{grade_row}"] = "Grade"
ws[f"B{grade_row}"] = "Count"
ws[f"C{grade_row}"] = "Percentage"
ws[f"A{grade_row}"].font = Font(bold=True)
ws[f"B{grade_row}"].font = Font(bold=True)
ws[f"C{grade_row}"].font = Font(bold=True)

for i, (grade, count) in enumerate(grade_distribution.items(), start=grade_row + 1):
    ws[f"A{i}"] = grade
    ws[f"B{i}"] = count
    ws[f"C{i}"] = f"{(count / total_students) * 100:.2f}%"

# Bar chart for grade distribution
bar_chart = BarChart()
bar_chart.title = "Grade Distribution"
bar_chart.x_axis.title = "Grades"
bar_chart.y_axis.title = "Number of Students"

bar_data = Reference(ws, min_col=2, min_row=grade_row, max_row=grade_row + len(grade_distribution))
bar_categories = Reference(ws, min_col=1, min_row=grade_row + 1, max_row=grade_row + len(grade_distribution))
bar_chart.add_data(bar_data, titles_from_data=True)
bar_chart.set_categories(bar_categories)
ws.add_chart(bar_chart, f"D{summary_start_row}")

# Pie chart for grade distribution
pie_chart = PieChart()
pie_chart.title = "Grade Distribution (Pie Chart)"
pie_chart.add_data(bar_data, titles_from_data=True)
pie_chart.set_categories(bar_categories)
ws.add_chart(pie_chart, f"D{summary_start_row + 15}")

# Histogram for Average Marks
histogram_data = Reference(ws, min_col=8, min_row=2, max_row=len(df) + 1)  # Average Marks column
histogram_chart = BarChart()
histogram_chart.title = "Histogram of Average Marks"
histogram_chart.add_data(histogram_data, titles_from_data=False)
ws.add_chart(histogram_chart, f"D{summary_start_row + 30}")

# Scatter plot for Math vs. Science scores
scatter_chart = ScatterChart()
scatter_chart.title = "Math vs. Science Scores"
x_values = Reference(ws, min_col=2, min_row=2, max_row=len(df) + 1)  # Math scores
y_values = Reference(ws, min_col=3, min_row=2, max_row=len(df) + 1)  # Science scores
series = Series(y_values, x_values, title="Math vs Science")
scatter_chart.series.append(series)
ws.add_chart(scatter_chart, f"D{summary_start_row + 45}")

# Line chart for Total Marks trend
line_chart = LineChart()
line_chart.title = "Total Marks Trend"
line_data = Reference(ws, min_col=7, min_row=2, max_row=len(df) + 1)  # Total Marks column
line_categories = Reference(ws, min_col=1, min_row=2, max_row=len(df) + 1)  # Student Names
line_chart.add_data(line_data, titles_from_data=False)
line_chart.set_categories(line_categories)
ws.add_chart(line_chart, f"D{summary_start_row + 60}")

# Format the columns
for col in ws.columns:
    for cell in col:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Save the workbook with the dashboard
wb.save(file_path)

print(f"Grades have been created and exported to {file_path}")
